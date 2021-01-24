import os
from openpyxl import load_workbook,Workbook
if os.path.exists("studentdetails.xlsx"):
    wb = load_workbook("studentdetails.xlsx")
    ws = wb.worksheets[0]
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["USN","Name","M1","M2","M3"])
n=int(input("Enter the number of records to be entered:"))
for i in range(n):
    print("Enter USN Name Marks1 Marks2 Marks3")
    data = input().split(" ")
    ws.append([data[0],data[1],data[2],data[3],data[4]])

wb.save("studentdetails.xlsx")
os.system("studentdetails.xlsx")
